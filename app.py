# app.py
import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Relatório de Faturamento & Premiações", layout="wide")
st.title("📊 Relatório de Faturamento & Premiações")

# =========================
# Helpers de formatação
# =========================
def format_moeda_str(x):
    try:
        return f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return ""

def format_percent_str(x):
    try:
        return f"{x:.1%}".replace(".", ",")
    except:
        return ""

def df_for_display(df, moeda_cols, percent_cols, cols_order=None):
    """Cria cópia do DF com strings de exibição (R$ e %) para o Streamlit."""
    df_disp = df.copy()
    if cols_order:
        df_disp = df_disp[cols_order]
    for c in moeda_cols:
        if c in df_disp.columns:
            df_disp[c] = df_disp[c].apply(lambda v: format_moeda_str(v) if pd.notnull(v) else "")
    for c in percent_cols:
        if c in df_disp.columns:
            df_disp[c] = df_disp[c].apply(lambda v: format_percent_str(v) if pd.notnull(v) else "")
    return df_disp

def to_excel_bytes_with_formats(df, sheet_name="Sheet1", moeda_cols=None, percent_cols=None, sombreamento_loja=True):
    """Gera bytes de um Excel com formatação (moeda/%) e larguras automáticas."""
    if moeda_cols is None:
        moeda_cols = []
    if percent_cols is None:
        percent_cols = []

    # Grava primeiro com pandas/openpyxl
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb[sheet_name]

    # Cabeçalho cinza e negrito
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Índices de colunas por nome
    header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

    # Sombreamento leve na coluna LOJA (linhas de dados)
    if sombreamento_loja and "LOJA" in header_map:
        loja_col = header_map["LOJA"]
        loja_letter = get_column_letter(loja_col)
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            ws[f"{loja_letter}{r}"].fill = light_gray

    # Formatos
    for nome in moeda_cols:
        if nome in header_map:
            cidx = header_map[nome]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                if cell.value not in (None, ""):
                    cell.number_format = 'R$ #,##0.00'

    for nome in percent_cols:
        if nome in header_map:
            cidx = header_map[nome]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                if cell.value not in (None, ""):
                    cell.number_format = '0.0%'

    # Largura automática
    for col in ws.columns:
        max_len = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max(12, max_len + 2)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# =========================
# Processamentos
# =========================
def process_faturamento(df_desvend: pd.DataFrame) -> pd.DataFrame:
    # Colunas mínimas
    required = ["LOJA", "VENDEDOR", "COTA TOTAL", "TOTAL VENDAS", "QUANT VENDAS", "SALDO COTA TOTAL"]
    faltando = [c for c in required if c not in df_desvend.columns]
    if faltando:
        raise ValueError(f"Colunas ausentes em DesVend: {', '.join(faltando)}")

    # Agrupar por LOJA + VENDEDOR
    grp = df_desvend.groupby(["LOJA", "VENDEDOR"], as_index=False).agg({
        "COTA TOTAL": "sum",
        "TOTAL VENDAS": "sum",
        "QUANT VENDAS": "sum",
        "SALDO COTA TOTAL": "sum"
    })

    # Métricas
    grp["% VENDAS"] = grp["TOTAL VENDAS"] / grp["COTA TOTAL"]
    grp["TICK MEDIO"] = grp.apply(lambda r: (r["TOTAL VENDAS"] / r["QUANT VENDAS"]) if r["QUANT VENDAS"] else 0.0, axis=1)
    grp["% SALDO COTA"] = grp["SALDO COTA TOTAL"] / grp["COTA TOTAL"]

    # Ordem final (inclui VENDEDOR após LOJA)
    ordem = ["LOJA", "VENDEDOR", "COTA TOTAL", "TOTAL VENDAS", "QUANT VENDAS",
             "% VENDAS", "TICK MEDIO", "SALDO COTA TOTAL", "% SALDO COTA"]
    grp = grp[ordem]

    # Ordenar visualmente por LOJA e VENDEDOR
    grp = grp.sort_values(["LOJA", "VENDEDOR"], kind="stable").reset_index(drop=True)
    return grp

def process_premiacoes(df_desvend: pd.DataFrame, df_taloes: pd.DataFrame,
                       pct_threshold: float, valor_premio: float) -> pd.DataFrame:
    # Garantir colunas
    if "VENDEDOR" not in df_taloes.columns:
        raise ValueError("A planilha Talões Pendentes precisa ter a coluna 'VENDEDOR'.")

    for c in ["VENDAS FORA DA POLÍTICA", "VENDAS ATUALIZADAS", "% VENDAS ATUALIZADAS"]:
        if c not in df_taloes.columns:
            df_taloes[c] = pd.NA

    # Merge por VENDEDOR (somente interseção para garantir correspondência)
    merged = pd.merge(
        df_desvend,
        df_taloes[["VENDEDOR", "VENDAS FORA DA POLÍTICA", "VENDAS ATUALIZADAS", "% VENDAS ATUALIZADAS"]],
        on="VENDEDOR",
        how="inner",
        suffixes=("", "_TALOES")
    )

    # Normalizações
    merged["VENDAS FORA DA POLÍTICA"] = pd.to_numeric(merged["VENDAS FORA DA POLÍTICA"], errors="coerce").fillna(0.0)
    # Se VENDAS ATUALIZADAS não vier, usar TOTAL VENDAS - VENDAS FORA...
    merged["VENDAS ATUALIZADAS"] = pd.to_numeric(merged["VENDAS ATUALIZADAS"], errors="coerce")
    merged["VENDAS ATUALIZADAS"] = merged.apply(
        lambda r: r["VENDAS ATUALIZADAS"] if pd.notnull(r["VENDAS ATUALIZADAS"]) else (r["TOTAL VENDAS"] - r["VENDAS FORA DA POLÍTICA"]),
        axis=1
    )
    # % VENDAS ATUALIZADAS: usar da planilha se veio; caso contrário, recalcular
    merged["% VENDAS ATUALIZADAS"] = pd.to_numeric(merged["% VENDAS ATUALIZADAS"], errors="coerce")
    merged["% VENDAS ATUALIZADAS"] = merged.apply(
        lambda r: r["% VENDAS ATUALIZADAS"] if pd.notnull(r["% VENDAS ATUALIZADAS"])
        else (r["VENDAS ATUALIZADAS"] / r["COTA TOTAL"] if r["COTA TOTAL"] else 0.0),
        axis=1
    )

    # Agregar por LOJA
    agg = merged.groupby("LOJA", as_index=False).agg({
        "COTA TOTAL": "sum",
        "TOTAL VENDAS": "sum",
        "SALDO COTA TOTAL": "sum",
        "VENDAS FORA DA POLÍTICA": "sum",
        "VENDAS ATUALIZADAS": "sum",
        "% VENDAS ATUALIZADAS": "mean"
    })

    # Derivadas
    agg["% VENDAS"] = agg.apply(lambda r: (r["TOTAL VENDAS"] / r["COTA TOTAL"]) if r["COTA TOTAL"] else 0.0, axis=1)
    agg["% SALDO COTA"] = agg.apply(lambda r: (r["SALDO COTA TOTAL"] / r["COTA TOTAL"]) if r["COTA TOTAL"] else 0.0, axis=1)

    # Regras de premiação (automático)
    agg["PREMIADO"] = agg["% VENDAS ATUALIZADAS"] >= pct_threshold
    agg["VALOR"] = agg["PREMIADO"].apply(lambda b: valor_premio if b else 0.0)
    agg["TOTAL LOJA"] = agg["VALOR"]  # total de prêmios por loja

    # Ordem final
    cols_order = ["LOJA", "COTA TOTAL", "TOTAL VENDAS", "% VENDAS", "SALDO COTA TOTAL", "% SALDO COTA",
                  "VENDAS FORA DA POLÍTICA", "VENDAS ATUALIZADAS", "% VENDAS ATUALIZADAS", "VALOR", "TOTAL LOJA"]
    cols_order = [c for c in cols_order if c in agg.columns]
    agg = agg[cols_order].sort_values("LOJA", kind="stable").reset_index(drop=True)
    return agg

# =========================
# UI - Abas
# =========================
tab_fat, tab_prem = st.tabs(["Faturamento", "Premiações"])

with tab_fat:
    st.subheader("📁 Upload — DesVend.xlsx")
    arquivo_desvend = st.file_uploader("Selecione o arquivo DesVend (xlsx). Se existir aba 'DesVend', ela será usada.", type=["xlsx"], key="fat1")

    if arquivo_desvend:
        # Ler planilha (tenta aba DesVend; senão, primeira aba)
        try:
            df_desvend = pd.read_excel(arquivo_desvend, sheet_name="DesVend")
        except Exception:
            df_desvend = pd.read_excel(arquivo_desvend)

        # Processar faturamento (LOJA + VENDEDOR)
        try:
            df_fat = process_faturamento(df_desvend)
        except Exception as e:
            st.error(f"Erro ao processar DesVend: {e}")
            st.stop()

        # Exibir tabela formatada no Streamlit
        moeda_cols_fat = ["COTA TOTAL", "TOTAL VENDAS", "TICK MEDIO", "SALDO COTA TOTAL"]
        percent_cols_fat = ["% VENDAS", "% SALDO COTA"]
        ordem_fat = ["LOJA", "VENDEDOR", "COTA TOTAL", "TOTAL VENDAS", "QUANT VENDAS",
                     "% VENDAS", "TICK MEDIO", "SALDO COTA TOTAL", "% SALDO COTA"]

        st.subheader("📊 Tabela Consolidada — por LOJA e VENDEDOR")
        st.dataframe(
            df_for_display(df_fat, moeda_cols_fat, percent_cols_fat, cols_order=ordem_fat),
            use_container_width=True,
            hide_index=True
        )

        # Gráfico: agregado por LOJA (mais harmonioso), largura da página, altura dinâmica conforme nº de lojas
        agg_loja = df_fat.groupby("LOJA", as_index=False)["TOTAL VENDAS"].sum().sort_values("TOTAL VENDAS", ascending=False)
        n_lojas = agg_loja.shape[0]
        height = max(450, 36 * n_lojas)  # altura cresce com número de lojas

        fig = px.bar(
            agg_loja,
            x="LOJA",
            y="TOTAL VENDAS",
            text="TOTAL VENDAS",
            title="Total de Vendas por Loja"
        )
        fig.update_traces(texttemplate='R$ %{y:,.2f}', textposition='outside')
        fig.update_layout(xaxis_tickangle=-35, height=height)
        st.plotly_chart(fig, use_container_width=True)

        # Download Excel com formatação real
        excel_bytes = to_excel_bytes_with_formats(
            df_fat,
            sheet_name="Faturamento",
            moeda_cols=moeda_cols_fat,
            percent_cols=percent_cols_fat
        )
        st.download_button(
            "📥 Baixar Excel — Faturamento",
            data=excel_bytes,
            file_name="Faturamento.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with tab_prem:
    st.subheader("📁 Upload — DesVend.xlsx e TALÕES PENDENTES.xlsx")
    arquivo_desvend2 = st.file_uploader("DesVend (xlsx)", type=["xlsx"], key="prem1")
    arquivo_taloes = st.file_uploader("Talões Pendentes (xlsx)", type=["xlsx"], key="prem2")

    col1, col2 = st.columns(2)
    with col1:
        pct_meta_view = st.number_input("Percentual mínimo para premiação (%)", min_value=0.0, max_value=100.0, value=45.0, step=0.5)
    with col2:
        valor_premio = st.number_input("Valor da premiação por loja (R$)", min_value=0.0, value=100.0, step=10.0)
    pct_meta = pct_meta_view / 100.0

    if arquivo_desvend2 and arquivo_taloes:
        # Ler planilhas
        try:
            try:
                df_desvend2 = pd.read_excel(arquivo_desvend2, sheet_name="DesVend")
            except Exception:
                df_desvend2 = pd.read_excel(arquivo_desvend2)

            df_taloes = pd.read_excel(arquivo_taloes)
        except Exception as e:
            st.error(f"Erro ao ler arquivos: {e}")
            st.stop()

        # Consolidado de premiações (por LOJA)
        try:
            df_prem = process_premiacoes(df_desvend2, df_taloes, pct_meta, valor_premio)
        except Exception as e:
            st.error(f"Erro ao processar Premiações: {e}")
            st.stop()

        moeda_cols_prem = ["COTA TOTAL", "TOTAL VENDAS", "SALDO COTA TOTAL", "VENDAS FORA DA POLÍTICA",
                           "VENDAS ATUALIZADAS", "VALOR", "TOTAL LOJA"]
        percent_cols_prem = ["% VENDAS", "% SALDO COTA", "% VENDAS ATUALIZADAS"]

        ordem_prem = ["LOJA", "COTA TOTAL", "TOTAL VENDAS", "% VENDAS", "SALDO COTA TOTAL", "% SALDO COTA",
                      "VENDAS FORA DA POLÍTICA", "VENDAS ATUALIZADAS", "% VENDAS ATUALIZADAS", "VALOR", "TOTAL LOJA"]

        st.subheader("🏆 Premiações — por LOJA")
        st.dataframe(
            df_for_display(df_prem, moeda_cols_prem, percent_cols_prem, cols_order=ordem_prem),
            use_container_width=True,
            hide_index=True
        )

        # Gráfico opcional de TOTAL LOJA (soma das premiações)
        fig2 = px.bar(
            df_prem.sort_values("TOTAL LOJA", ascending=False),
            x="LOJA",
            y="TOTAL LOJA",
            text="TOTAL LOJA",
            title="Total de Premiações por Loja"
        )
        fig2.update_traces(texttemplate='R$ %{y:,.2f}', textposition='outside')
        fig2.update_layout(xaxis_tickangle=-35, height=max(450, 36 * df_prem.shape[0]))
        st.plotly_chart(fig2, use_container_width=True)

        # Excel com formatação real
        excel_prem = to_excel_bytes_with_formats(
            df_prem,
            sheet_name="Premiações",
            moeda_cols=moeda_cols_prem,
            percent_cols=percent_cols_prem
        )
        st.download_button(
            "📥 Baixar Excel — Premiações",
            data=excel_prem,
            file_name="Premiacoes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
